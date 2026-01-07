import requests
import pandas as pd
import time
import os
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm

# === Konfiguracja ===
INPUT_FILE = "bajki.xlsx"
OUTPUT_FILE = "bajki-done.xlsx"
BACKUP_INTERVAL = 5  # Częstszy zapis, bo Jikan jest wolniejszy

# === AniList Config ===
ANILIST_API_URL = "https://graphql.anilist.co"

# === Jikan (MAL) Config ===
JIKAN_API_URL = "https://api.jikan.moe/v4"

# === Logowanie ===
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# --- CZĘŚĆ 1: ANILIST ---

ANILIST_QUERY = """
query($name: String) {
  Media(search: $name, type: ANIME, sort: SEARCH_MATCH) {
    title { romaji english }
    season
    seasonYear
    startDate { year month }
    siteUrl
    studios(sort: MAIN) {
      edges { isMain node { name } }
    }
    staff(perPage: 1, sort: RELEVANCE) {
      edges { role node { name { full } } }
    }
    genres
    tags { name rank }
  }
}
"""


def fetch_anilist(name):
    """Pobiera dane z AniList."""
    try:
        response = requests.post(
            ANILIST_API_URL,
            json={"query": ANILIST_QUERY, "variables": {"name": name}},
            timeout=5
        )

        # Rate Limiting
        remaining = int(response.headers.get('X-RateLimit-Remaining', 0))
        if response.status_code == 429:
            time.sleep(60)  # Kara za spam
            return None

        if response.status_code == 200:
            return response.json().get("data", {}).get("Media")

    except Exception:
        pass
    return None


def parse_anilist(media):
    """Parsuje dane z AniList na format wyjściowy."""
    if not media: return None

    # Sezon
    season = media.get("season")
    year = media.get("seasonYear")
    season_str = f"{season} {year}" if season and year else None

    if not season_str:
        s_date = media.get("startDate", {})
        y, m = s_date.get("year"), s_date.get("month")
        season_str = calculate_fallback_season(y, m)

    # Studio / Staff
    studio_str = "Unknown"
    studios = media.get("studios", {}).get("edges", [])
    if studios:
        studio_str = studios[0]["node"]["name"]
    else:
        # Fallback do Staff
        staff = media.get("staff", {}).get("edges", [])
        if staff:
            p = staff[0]
            studio_str = f"{p['node']['name']['full']} ({p['role']})"

    genres = ", ".join(media.get("genres", []))
    tags = ", ".join([t["name"] for t in media.get("tags", []) if t.get("rank", 0) > 40][:5])
    link = media.get("siteUrl", "")

    return season_str, studio_str, genres, tags, link


# --- CZĘŚĆ 2: JIKAN (MYANIMELIST) ---

def fetch_jikan(name):
    """Pobiera dane z Jikan (MAL) z opóźnieniem."""
    time.sleep(1.5)  # Jikan wymaga ~1s przerwy między zapytaniami
    url = f"{JIKAN_API_URL}/anime"
    params = {"q": name, "limit": 1}

    try:
        response = requests.get(url, params=params, timeout=10)
        if response.status_code == 200:
            data = response.json().get("data", [])
            return data[0] if data else None
        elif response.status_code == 429:
            time.sleep(5)  # Krótka przerwa przy limicie
    except Exception as e:
        logger.warning(f"Jikan error dla {name}: {e}")
    return None


def fetch_jikan_staff(mal_id):
    """Pobiera staff z MAL, jeśli brak studia."""
    time.sleep(1.5)
    url = f"{JIKAN_API_URL}/anime/{mal_id}/staff"
    try:
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            data = response.json().get("data", [])
            # Szukamy reżysera lub autora
            for person in data:
                roles = person.get("positions", [])
                if any("Director" in r or "Original Creator" in r for r in roles):
                    return f"{person['person']['name']} ({roles[0]})"
            # Jeśli nie ma reżysera, bierz pierwszego z brzegu
            if data:
                return f"{data[0]['person']['name']} ({data[0]['positions'][0]})"
    except Exception:
        pass
    return "Unknown"


def parse_jikan(data):
    """Parsuje dane z Jikan (MAL)."""
    if not data: return None

    # Sezon
    season = data.get("season")
    year = data.get("year")
    season_str = f"{season.upper()} {year}" if season and year else None

    if not season_str:
        aired = data.get("aired", {}).get("prop", {}).get("from", {})
        y, m = aired.get("year"), aired.get("month")
        season_str = calculate_fallback_season(y, m)

    # Studio / Staff
    studios = data.get("studios", [])
    if studios:
        studio_str = studios[0]["name"]
    else:
        # Pobieramy staff osobnym zapytaniem
        studio_str = fetch_jikan_staff(data["mal_id"])

    genres_list = [g["name"] for g in data.get("genres", [])]
    themes_list = [t["name"] for t in data.get("themes", [])]

    genres = ", ".join(genres_list)
    tags = ", ".join(themes_list[:5])  # MAL używa themes jako tagów
    link = data.get("url", "")

    return season_str, studio_str, genres, tags, link


# --- LOGIKA WSPÓLNA ---

def calculate_fallback_season(year, month):
    """Uniwersalna logika 'Winter + Rok'."""
    if not year: return "Unknown"

    if month:
        if 1 <= month <= 3: return f"WINTER {year}"
        if 4 <= month <= 6: return f"SPRING {year}"
        if 7 <= month <= 9: return f"SUMMER {year}"
        return f"FALL {year}"

    return f"WINTER {year}"


def process_title(title):
    """Główna funkcja decyzyjna."""

    # 1. Próba AniList
    al_data = fetch_anilist(title)
    result = parse_anilist(al_data)

    if result:
        return result

    # 2. Próba Jikan (MAL)
    logger.info(f"🔄 AniList pusty dla '{title}', próbuję MAL...")
    mal_data = fetch_jikan(title)
    result = parse_jikan(mal_data)

    if result:
        return result

    return ["Unknown"] * 5


# --- OBSŁUGA PLIKÓW ---

def main():
    if not os.path.exists(INPUT_FILE):
        print("Brak pliku wejściowego.")
        return

    # Wczytanie / Wznawianie
    df = pd.read_excel(INPUT_FILE)
    if "Bajka" not in df.columns:
        df.rename(columns={df.columns[0]: "Bajka"}, inplace=True)

    cols = ["Sezon", "Studio", "Gatunki", "Tagi", "Link"]
    for col in cols:
        if col not in df.columns: df[col] = None

    if os.path.exists(OUTPUT_FILE):
        print("📂 Wczytuję istniejące wyniki...")
        df_done = pd.read_excel(OUTPUT_FILE)
        df = df.set_index("Bajka")
        df.update(df_done.set_index("Bajka"))
        df.reset_index(inplace=True)

    # Lista do zrobienia
    mask = df["Link"].isnull() | (df["Link"] == "") | (df["Link"] == "Unknown")
    todo_idx = df[mask].index.tolist()

    print(f"🚀 Do pobrania: {len(todo_idx)} pozycji.")

    try:
        for i, idx in enumerate(tqdm(todo_idx, unit="tytuł")):
            title = df.at[idx, "Bajka"]

            # Pobranie danych (AniList -> Fallback MAL)
            data = process_title(title)

            df.loc[idx, cols] = data

            # Auto-save
            if (i + 1) % BACKUP_INTERVAL == 0:
                df.to_excel(OUTPUT_FILE, index=False)

    except KeyboardInterrupt:
        print("\n🛑 Przerwano.")
    finally:
        df.to_excel(OUTPUT_FILE, index=False)
        highlight_unknowns(OUTPUT_FILE)
        print("✅ Zakończono.")


def highlight_unknowns(path):
    try:
        wb = load_workbook(path)
        ws = wb.active
        fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        for row in ws.iter_rows(min_row=2):
            # Jeśli link pusty lub Unknown -> koloruj
            if row[4].value in [None, "", "Unknown"]:
                for cell in row: cell.fill = fill
        wb.save(path)
    except:
        pass


if __name__ == "__main__":
    main()
